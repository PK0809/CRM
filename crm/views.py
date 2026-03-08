from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout, get_user_model
from django.contrib.auth.models import Permission
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse, HttpResponseForbidden
from django.db import IntegrityError, transaction
from django.urls import reverse_lazy
from django.views.generic import UpdateView
from django.utils.timezone import now, localdate
from django.db.models import Count, Sum, Q
from django.core.paginator import Paginator
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST, require_http_methods
from django.template.loader import render_to_string
from django.conf import settings
from django.views.decorators.csrf import csrf_protect
from django.contrib.auth.decorators import user_passes_test
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Permission
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_protect
from django.contrib import messages
from django.shortcuts import render, redirect

from .models import UserProfile


from decimal import Decimal, InvalidOperation
from datetime import timedelta
from pathlib import Path

from .models import (
    UserProfile, Client, Invoice, Lead, Estimation, EstimationItem, Branch,
    UserPermission, PaymentLog, GSTSettings, EstimationSettings,
    TermsAndConditions, DeliveryChallan, DeliveryChallanItem
)
from .forms import UserForm, ClientForm, EstimationForm, ApprovalForm
from .utils import inr_currency_words, generate_invoice_number

User = get_user_model()

from django.contrib.auth import authenticate, login

def user_login(request):
    if request.method == "POST":
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            request.user.refresh_from_db()  # ✅ Refresh permissions
            return redirect('dashboard')
        else:
            messages.error(request, "Invalid username or password.")
    return render(request, 'auth/login.html')

def logout_view(request):
    logout(request)
    return redirect('login')


# ---------------------------------------------------
# User Management Views
# ---------------------------------------------------
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.urls import reverse_lazy
from django.views.generic import UpdateView, ListView, DeleteView
from django.http import JsonResponse
from django.contrib.auth import update_session_auth_hash 

from .forms import UserForm
from .models import User, UserProfile, UserPermission



# ---------------------------------------------------
# Create User
# ---------------------------------------------------
@login_required(login_url='login')
@csrf_protect
def create_user(request):
    if request.method == "POST":
        username = request.POST.get("username")
        email = request.POST.get("email")
        password = request.POST.get("password")
        confirm_password = request.POST.get("confirm_password")
        role = request.POST.get("role")
        phone_number = request.POST.get("phone_number")
        permissions = request.POST.getlist("permissions")

        if password != confirm_password:
            messages.error(request, "Passwords do not match.")
            return redirect("create_user")

        if User.objects.filter(username=username).exists():
            messages.error(request, "Username already exists.")
            return redirect("create_user")

        user = User.objects.create_user(username=username, email=email, password=password)
        user.is_staff = (role == "Admin")
        user.save()

        # ✅ Assign permissions safely
        if permissions:
            codenames = [p.split('.')[-1] for p in permissions]
            perms = Permission.objects.filter(codename__in=codenames)
            user.user_permissions.set(perms)

        # ✅ Safe profile creation
        UserProfile.objects.update_or_create(
            user=user,
            defaults={
                "name": username,
                "phone_number": phone_number,
                "role": role,
            }
        )

        messages.success(request, f"User '{username}' created successfully!")
        return redirect("user_list")

    permissions = Permission.objects.filter(content_type__app_label='crm')
    return render(request, "users/add_user.html", {"permissions": permissions})



# ---------------------------------------------------
# User List
# ---------------------------------------------------
@login_required(login_url='login')
def user_list(request):
    """Display all user profiles (admin only)."""
    users = UserProfile.objects.select_related("user").all().order_by("-id")
    return render(request, "users/user_list.html", {"users": users})

    
@login_required(login_url='login')
def edit_user(request, pk):
    user = get_object_or_404(User, pk=pk)

    if request.method == 'POST':
        form = UserForm(request.POST, instance=user)
        if form.is_valid():
            form.save()

            # ✅ Important: keep the user logged in after password change
            update_session_auth_hash(request, user)

            messages.success(request, f"✅ User '{user.username}' updated successfully.")
            return redirect('user_list')
        else:
            messages.error(request, "⚠️ Please correct the errors below.")
    else:
        form = UserForm(instance=user)

    return render(request, 'users/user_form.html', {'form': form, 'user': user})



from django.contrib import messages
from django.shortcuts import redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.decorators.csrf import csrf_protect
from django.contrib.auth import get_user_model
from .models import UserProfile

User = get_user_model()


def is_admin(user):
    return user.is_superuser or (
        hasattr(user, 'userprofile') and user.userprofile.role == 'Admin'
    )



@login_required(login_url='login')
@user_passes_test(is_admin, login_url='login')
@csrf_protect
def user_delete(request, pk):
    user = get_object_or_404(User, pk=pk)

    if request.method == 'POST':
        username = user.username
        user.delete()
        messages.success(request, f"User '{username}' deleted successfully.")
        return redirect('user_list')

    return render(request, 'users/user_delete.html', {'user': user})



# ---------------------------------------------------
# Get Permissions by Role (AJAX)
# ---------------------------------------------------
@login_required
def get_permissions_by_role(request):
    """AJAX helper: return permissions for selected role."""
    role = request.GET.get("role")
    permissions = list(UserPermission.objects.values("id", "name").order_by("name"))
    if role == "Admin":
        permissions = []  # Admins have implicit full rights
    return JsonResponse({"permissions": permissions})


# ---------------------------------------------------
# Logout View
# ---------------------------------------------------
@login_required
def logout_view(request):
    logout(request)
    messages.info(request, "You have been logged out.")
    return redirect("login")

from django.views.generic import UpdateView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.urls import reverse_lazy
from django.contrib import messages
from django.contrib.auth import get_user_model
from .forms import UserForm

User = get_user_model()

class UserUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = User
    form_class = UserForm
    template_name = 'users/user_form.html'
    success_url = reverse_lazy('user_list')

    def test_func(self):
        """Allow only Admins and Superusers to edit users"""
        user = self.request.user
        return user.is_superuser or (hasattr(user, 'userprofile') and user.userprofile.role == 'Admin')

    def handle_no_permission(self):
        messages.error(self.request, "You don’t have permission to edit users.")
        return redirect('user_list')

    def form_valid(self, form):
        """Optional: display success message"""
        messages.success(self.request, f"User '{form.instance.username}' updated successfully.")
        return super().form_valid(form)


from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.db.models import Sum, Count, Q
from .models import Invoice, PaymentLog, Lead, Estimation, Client, UserPermission


@login_required
def dashboard(request):
    user = request.user

    # --- USER PERMISSIONS ---
    user_perm_names = set(
        UserPermission.objects
        .filter(userprofile__user=user)
        .values_list("name", flat=True)
    )

    context = {
        "can_view_client": "can_view_client" in user_perm_names,
        "can_add_client": "can_add_client" in user_perm_names,
        "can_view_lead": "can_view_lead" in user_perm_names,
        "can_view_estimation": "can_view_estimation" in user_perm_names,
        "can_view_invoice": "can_view_invoice" in user_perm_names,
        "can_view_reports": "can_view_reports" in user_perm_names,
        "grouped_modules": {
            "Sales": [
                {"name": "Client", "url": "/client/"},
                {"name": "Lead", "url": "/lead/"},
                {"name": "Estimation", "url": "/estimation/"},
                {"name": "Delivery Challan", "url": "/dc/"},
                {"name": "Invoice", "url": "/invoices/"},
                {"name": "Reports", "url": "/reports/"},
            ],
            "Purchase": [
                {"name": "Vendor", "url": "/vendor/"},
                {"name": "PO", "url": "/purchase-order/"},
                {"name": "Bill", "url": "/bill/"},
            ],
        },
    }

    # --- AGGREGATES ---
    context["total_invoiced"] = (
        Invoice.objects.aggregate(total=Sum("total_value"))["total"] or 0
    )

    context["paid"] = (
        PaymentLog.objects.filter(
            Q(status="Paid") | Q(status="Partial Paid")
        ).aggregate(total=Sum("amount_paid"))["total"] or 0
    )

    context["balance_due"] = (
        Invoice.objects.aggregate(total=Sum("balance_due"))["total"] or 0
    )

    # --- COUNTS ---
    total_leads = Lead.objects.count()
    total_quotations = Estimation.objects.count()
    total_invoices = Invoice.objects.count()

    context.update({
        "total_leads": total_leads,
        "total_quotations": total_quotations,
        "total_invoices": total_invoices,
        "conversion_rate": round((total_invoices / total_leads) * 100, 2)
        if total_leads else 0,
    })

    # --- STATUS BREAKDOWNS (✅ JSON SAFE) ---
    quotation_status_qs = (
        Estimation.objects
        .exclude(status__isnull=True)
        .values("status")
        .annotate(count=Count("id"))
        .order_by("status")
    )

    context["quotation_status"] = [
        {"status": row["status"], "count": row["count"]}
        for row in quotation_status_qs
    ]

    invoice_status_qs = (
        Invoice.objects
        .exclude(status__isnull=True)
        .values("status")
        .annotate(count=Count("id"))
        .order_by("status")
    )

    context["invoice_status"] = [
        {"status": row["status"], "count": row["count"]}
        for row in invoice_status_qs
    ]

    # --- TOP CLIENTS (NOT USED IN JSON → SAFE AS QS) ---
    context["top_clients"] = (
        Client.objects
        .annotate(total_leads=Count("lead"))
        .order_by("-total_leads")[:4]
    )

    # --- FILTER OPTIONS ---
    context["filter_options"] = [
        ("This Month", "this_month"),
        ("This Quarter", "this_quarter"),
        ("This Year", "this_year"),
        ("Previous Month", "previous_month"),
        ("Previous Quarter", "previous_quarter"),
        ("Previous Year", "previous_year"),
        ("Custom", "custom"),
    ]

    context["selected_filter"] = request.GET.get("date_filter", "this_month")
    context["user_name"] = user.first_name or user.username

    return render(request, "dashboard.html", context)


@login_required
def confirm_payment(request, payment_id):
    """Confirm a payment and update invoice status accordingly."""
    payment = get_object_or_404(PaymentLog, id=payment_id)
    invoice = payment.invoice

    # Total paid for this invoice
    total_paid = (
        PaymentLog.objects.filter(
            invoice=invoice, status__in=["Paid", "Partial Paid"]
        ).aggregate(total=Sum("amount_paid"))["total"]
        or 0
    )

    # --- UPDATE INVOICE STATUS ---
    if total_paid >= invoice.total_value:
        invoice.status = "Paid"
    elif total_paid > 0:
        invoice.status = "Partial Paid"
    else:
        invoice.status = "Unpaid"

    invoice.paid_amount = total_paid
    invoice.balance_due = invoice.total_value - total_paid
    invoice.save()

    # --- UPDATE PAYMENT LOG ---
    payment.status = invoice.status
    payment.save()

    return redirect("payment_list")



@login_required
def client_list(request):
    # 🔍 Get search query (trim spaces)
    query = request.GET.get('q', '').strip()

    # 📋 Get all clients
    all_clients = Client.objects.all().order_by('-id')
    total_clients = all_clients.count()

    # 🔎 Filter if search query is provided
    if query:
        clients = all_clients.filter(company_name__icontains=query).order_by('-id')
        filtered_count = clients.count()
    else:
        clients = all_clients
        filtered_count = total_clients

    # 📄 Pagination — 10 per page
    paginator = Paginator(clients, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # 🧭 Context for template
    context = {
        'clients': page_obj,   # paginated list
        'query': query,        # current search text
        'page_obj': page_obj,  # pagination object
        'total_clients': total_clients,
        'filtered_count': filtered_count,
    }

    return render(request, 'client.html', context)



from django.contrib import messages
from django.shortcuts import redirect, render
from django.utils.timezone import now
from django.contrib.auth.decorators import login_required
from .models import Client, Branch


@login_required
def client_entry(request):
    """Handles adding a new client and automatically creates a default branch."""
    if request.method == "POST":
        try:
            company_name = request.POST.get("company_name", "").strip()
            type_of_company = request.POST.get("type_of_company", "").strip()
            gst_no = request.POST.get("gst_no", "").strip()
            contact_person = request.POST.get("contact_person", "").strip()
            email = request.POST.get("email", "").strip()
            mobile = request.POST.get("mobile", "").strip()
            address = request.POST.get("address", "").strip()

            # Basic validation
            if not company_name or not mobile or not address:
                messages.error(request, "⚠️ Company Name, Mobile, and Address are required.")
                return redirect("client")

            # Create Client
            client = Client.objects.create(
                company_name=company_name,
                type_of_company=type_of_company,
                gst_no=gst_no,
                contact_person=contact_person,
                email=email,
                mobile=mobile,
                address=address,
                created_at=now(),
            )

            # ✅ Automatically create Primary Branch
            Branch.objects.create(
                client=client,
                branch_name="Primary",
                contact_person=contact_person,
                mobile=mobile,
                email=email,
                gst_no=gst_no,
                address=address,
            )

            messages.success(request, f"✅ Client '{client.company_name}' added successfully with default 'Primary' branch.")
            return redirect("client")

        except Exception as e:
            messages.error(request, f"❌ Error adding client: {e}")
            return redirect("client")

    return render(request, "client.html")
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from .models import Client, Branch


@login_required
@csrf_exempt
def client_entry_ajax(request):
    """Handles AJAX client creation and auto-adds a Primary branch."""
    if request.method == "POST":
        try:
            client = Client.objects.create(
                company_name=request.POST.get("company_name", "").strip(),
                type_of_company=request.POST.get("type_of_company", "").strip(),
                gst_no=request.POST.get("gst_no", "").strip(),
                contact_person=request.POST.get("contact_person", "").strip(),
                email=request.POST.get("email", "").strip(),
                mobile=request.POST.get("mobile", "").strip(),
                address=request.POST.get("address", "").strip(),
            )

            # ✅ Automatically create a Primary branch
            Branch.objects.create(
                client=client,
                branch_name="Primary",
                contact_person=client.contact_person,
                mobile=client.mobile,
                email=client.email,
                gst_no=client.gst_no,
                address=client.address,
            )

            return JsonResponse({
                "success": True,
                "client": {
                    "id": client.id,
                    "company_name": client.company_name
                }
            })

        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})

    return JsonResponse({"success": False, "error": "Invalid request"})



# ✅ Edit an existing client
@login_required
def edit_client(request, client_id):
    client = get_object_or_404(Client, id=client_id)

    if request.method == 'POST':
        client.company_name = request.POST.get('company_name', '').strip()
        client.type_of_company = request.POST.get('type_of_company', '').strip()
        client.gst_no = request.POST.get('gst_no', '').strip()
        client.contact_person = request.POST.get('contact_person', '').strip()
        client.mobile = request.POST.get('mobile', '').strip()
        client.email = request.POST.get('email', '').strip()
        client.address = request.POST.get('address', '').strip()
        
        # Auto-clear fields for Individual
        if client.type_of_company.lower() == 'individual':
            client.gst_no = ''
            client.contact_person = ''

        client.save()
        messages.success(request, f'Client "{client.company_name}" updated successfully.')
        return redirect('client')

    return render(request, 'clients/edit_client.html', {'client': client})


# ✅ Delete client confirmation
@login_required
def delete_client(request, client_id):
    client = get_object_or_404(Client, id=client_id)
    if request.method == 'POST':
        name = client.company_name
        client.delete()
        messages.success(request, f'Client "{name}" deleted successfully.')
        return redirect('client')
    return render(request, 'clients/delete_client.html', {'client': client})


@csrf_exempt
@login_required
def add_branch(request):
    """Add a new branch to a specific client."""
    if request.method == 'POST':
        try:
            client_id = request.POST.get('client_id')
            branch_name = request.POST.get('branch_name', '').strip()
            contact_person = request.POST.get('contact_person', '').strip()
            mobile = request.POST.get('mobile', '').strip()
            email = request.POST.get('email', '').strip()
            gst_no = request.POST.get('gst_no', '').strip()
            address = request.POST.get('address', '').strip()

            # ✅ Validation
            if not client_id or not branch_name or not mobile or not address:
                messages.error(request, "⚠️ All required fields must be filled.")
                return redirect('branch_list', client_id=client_id)

            # ✅ Get Client (with validation)
            client = get_object_or_404(Client, id=client_id)

            # ✅ Default Email if blank
            if not email:
                email = client.email or ""

            # ✅ Create Branch
            Branch.objects.create(
                client=client,
                branch_name=branch_name,
                contact_person=contact_person,
                mobile=mobile,
                email=email,
                gst_no=gst_no,
                address=address,
            )

            messages.success(request, f"✅ Branch '{branch_name}' added successfully.")
            return redirect('branch_list', client_id=client.id)

        except Exception as e:
            messages.error(request, f"❌ Error adding branch: {e}")
            return redirect('branch_list', client_id=request.POST.get('client_id'))

    # Non-POST request
    return JsonResponse({'success': False, 'error': 'Invalid request method.'})

from django.http import JsonResponse

@login_required
def get_client_info(request):
    """Return client details (email, contact_person, mobile) as JSON for auto-fill in Add Branch."""
    client_id = request.GET.get("client_id")
    if not client_id:
        return JsonResponse({"success": False, "error": "Missing client_id"})

    try:
        client = Client.objects.get(id=client_id)
        data = {
            "success": True,
            "email": client.email or "",
            "contact_person": client.contact_person or "",
            "mobile": client.mobile or "",
            "gst_no": client.gst_no or "",
            "address": client.address or "",
        }
        return JsonResponse(data)
    except Client.DoesNotExist:
        return JsonResponse({"success": False, "error": "Client not found"})

@csrf_exempt
@login_required
def edit_branch(request, branch_id):
    """Edit existing branch details (with email fix)."""
    branch = get_object_or_404(Branch, id=branch_id)

    if request.method == 'POST':
        try:
            branch.branch_name = request.POST.get('branch_name', '').strip() or branch.branch_name
            branch.contact_person = request.POST.get('contact_person', '').strip() or branch.contact_person
            branch.mobile = request.POST.get('mobile', '').strip() or branch.mobile
            branch.email = request.POST.get('email', '').strip()  # ✅ make sure it's saved
            branch.gst_no = request.POST.get('gst_no', '').strip() or branch.gst_no
            branch.address = request.POST.get('address', '').strip() or branch.address

            branch.save(update_fields=['branch_name', 'contact_person', 'mobile', 'email', 'gst_no', 'address'])  # ✅ enforce save

            messages.success(request, f"✅ Branch '{branch.branch_name}' updated successfully.")
            return redirect('branch_list', client_id=branch.client.id)

        except Exception as e:
            messages.error(request, f"❌ Error updating branch: {e}")
            return redirect('branch_list', client_id=branch.client.id)

    return JsonResponse({'success': False, 'error': 'Invalid request method'})

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .models import Client, Branch
from .models import Lead, Client, Branch

@login_required
def branch_list(request, client_id):
    """List, add, edit, and delete branches for a specific client."""
    client = get_object_or_404(Client, id=client_id)
    branches = Branch.objects.filter(client=client).order_by('-id')

    # Handle Add or Edit
    if request.method == 'POST':
        branch_id = request.POST.get('branch_id')
        branch_name = request.POST.get('branch_name', '').strip()
        contact_person = request.POST.get('contact_person', '').strip()
        mobile = request.POST.get('mobile', '').strip()
        gst_no = request.POST.get('gst_no', '').strip()
        address = request.POST.get('address', '').strip()

        if not branch_name or not mobile or not address:
            messages.error(request, "Branch Name, Mobile, and Address are required.")
        else:
            if branch_id:  # Edit existing
                branch = get_object_or_404(Branch, id=branch_id, client=client)
                branch.branch_name = branch_name
                branch.contact_person = contact_person
                branch.mobile = mobile
                branch.gst_no = gst_no
                branch.address = address
                branch.save()
                messages.success(request, f"✅ Branch '{branch.branch_name}' updated successfully.")
            else:  # Add new
                Branch.objects.create(
                    client=client,
                    branch_name=branch_name,
                    contact_person=contact_person,
                    mobile=mobile,
                    gst_no=gst_no,
                    address=address
                )
                messages.success(request, f"✅ Branch '{branch_name}' added successfully.")
            return redirect('branch_list', client_id=client.id)

    context = {'client': client, 'branches': branches}
    return render(request, 'branches/branch_list.html', context)


@login_required
def delete_branch(request, client_id, branch_id):
    branch = get_object_or_404(Branch, id=branch_id, client_id=client_id)
    branch_name = branch.branch_name
    branch.delete()
    messages.success(request, f"🗑️ Branch '{branch_name}' deleted successfully.")
    return redirect('branch_list', client_id=client_id)



from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Exists, OuterRef

@login_required
def lead_list(request):
    search_query = request.GET.get('q', '')

    # 🔹 Subquery: check if estimation exists for this lead
    estimation_exists = Estimation.objects.filter(
        lead_no=OuterRef('pk')
    )

    leads = (
        Lead.objects
        .annotate(has_estimation=Exists(estimation_exists))
        .order_by('-id')
    )

    if search_query:
        leads = leads.filter(company_name__company_name__icontains=search_query)

    paginator = Paginator(leads, 10)
    page_obj = paginator.get_page(request.GET.get('page'))

    return render(request, 'lead.html', {
        'leads': page_obj,
        'page_obj': page_obj,
        'clients': Client.objects.all(),
        'query': search_query,
    })


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.utils.timezone import now
from .models import Client, Branch, Lead


@login_required
@csrf_exempt
def lead_create(request):
    """
    Create a new Lead with full branch integration.
    - Pulls contact, email, mobile, address from Branch (if selected) or Client fallback.
    - Includes Lead Type.
    """
    clients = Client.objects.all().order_by("company_name")

    if request.method == "POST":
        try:
            client_id = request.POST.get("company_name")
            branch_id = request.POST.get("contact_person")  # branch dropdown value
            email = request.POST.get("email", "").strip()
            mobile = request.POST.get("mobile", "").strip()
            address = request.POST.get("address", "").strip()
            requirement = request.POST.get("requirement", "").strip()
            lead_type = request.POST.get("lead_type", "Referrals").strip()

            # ✅ Validation
            if not client_id:
                messages.error(request, "⚠️ Please select a company.")
                return redirect("lead_list")

            company = get_object_or_404(Client, id=client_id)

            # Initialize contact details
            contact_person_name = ""
            branch = None

            # ✅ If a branch is selected
            if branch_id and branch_id.isdigit():
                branch = Branch.objects.filter(id=branch_id, client=company).first()

                if branch:
                    contact_person_name = branch.contact_person or ""
                    email = email or branch.email or ""
                    mobile = mobile or branch.mobile or ""
                    address = address or branch.address

            # ✅ If no branch or missing data → fallback to client details
            if not branch:
                contact_person_name = company.contact_person or ""
                email = email or company.email or ""
                mobile = mobile or company.mobile or ""
                address = address or company.address or ""

            if not mobile:
                messages.error(request, "⚠️ Mobile number is required.")
                return redirect("lead_list")

            # ✅ Create Lead
            Lead.objects.create(
                company_name=company,
                contact_person=contact_person_name,
                email=email,
                mobile=mobile,
                address=address,
                requirement=requirement,
                lead_type=lead_type,
                status="Pending",
                computed_status="Pending",
                date=now().date(),
            )

            messages.success(request, f"✅ Lead created successfully for {company.company_name}.")
            return redirect("lead_list")

        except Exception as e:
            messages.error(request, f"❌ Error creating lead: {str(e)}")
            return redirect("lead_list")

    # Render the form page if not POST
    return render(request, "leads/lead_create.html", {"clients": clients})


@login_required
def lead_edit(request, pk):
    """
    Edit an existing Lead safely:
    - Uses client ID instead of name to prevent 'expected number' errors.
    - Resolves branch contact name correctly.
    - Auto-fills email, mobile, address from branch or client.
    - Protects 'Won' leads from being modified.
    """
    lead = get_object_or_404(Lead, pk=pk)

    # Prevent editing of 'Won' leads
    if lead.status == "Won":
        messages.warning(request, "⚠️ You cannot edit a lead with status 'Won'.")
        return redirect("lead_list")

    clients = Client.objects.all().order_by("company_name")
    branches = Branch.objects.filter(client=lead.company_name)

    if request.method == "POST":
        try:
            # Fetch posted data
            client_value = request.POST.get("company_name")
            branch_id = request.POST.get("contact_person")
            email = request.POST.get("email", "").strip()
            mobile = request.POST.get("mobile", "").strip()
            address = request.POST.get("address", "").strip()
            requirement = request.POST.get("requirement", "").strip()
            lead_type = request.POST.get("lead_type", "Referrals").strip()

            # 🧩 FIX: Get correct Client instance (ID or Name fallback)
            try:
                company = Client.objects.get(id=int(client_value))
            except (ValueError, Client.DoesNotExist):
                company = Client.objects.filter(company_name=client_value).first()

            if not company:
                messages.error(request, "⚠️ Invalid company selected.")
                return redirect("lead_edit", pk=lead.id)

            # Determine contact person name (branch-based or fallback)
            contact_person_name = ""
            if branch_id and branch_id.isdigit():
                branch = Branch.objects.filter(id=branch_id).first()
                if branch:
                    contact_person_name = branch.contact_person or ""
                    # Auto-fill missing fields from branch
                    email = email or getattr(branch, "email", "")
                    mobile = mobile or getattr(branch, "mobile", "")
                    address = address or getattr(branch, "address", "")
            else:
                # Fallback to client data
                contact_person_name = company.contact_person or ""
                email = email or company.email or ""
                mobile = mobile or company.mobile or ""
                address = address or company.address or ""

            # ✅ Update lead safely
            lead.company_name = company
            lead.contact_person = contact_person_name
            lead.email = email
            lead.mobile = mobile
            lead.address = address
            lead.requirement = requirement
            lead.lead_type = lead_type
            lead.computed_status = lead.status  # keep sync
            lead.save()

            messages.success(request, "✅ Lead updated successfully.")
            return redirect("lead_list")

        except Exception as e:
            messages.error(request, f"❌ Error updating lead: {e}")
            return redirect("lead_edit", pk=lead.id)

    # Context for template rendering
    context = {
        "lead": lead,
        "clients": clients,
        "branches": branches,
        "lead_type_options": [
            "Referrals",
            "E-mail",
            "Advertisements",
            "Website",
            "JD",
            "Social media",
        ],
    }

    return render(request, "edit_lead.html", context)


def get_pending_lead(request):
    client_id = request.GET.get('client_id')
    lead = Lead.objects.filter(company_name_id=client_id, status="Pending").order_by('-date').first()
    if lead:
        return JsonResponse({'lead_no': lead.lead_no})
    return JsonResponse({'lead_no': ''})

def get_pending_leads(request):
    client_id = request.GET.get('client_id')
    leads = Lead.objects.filter(company_name_id=client_id, status='Pending')
    return JsonResponse({'leads': [{'id': lead.id, 'lead_no': lead.lead_no} for lead in leads]})

def get_gst_no(request):
    client_id = request.GET.get('client_id')
    try:
        client = Client.objects.get(id=client_id)
        return JsonResponse({'gst_no': client.gst_no})
    except Client.DoesNotExist:
        return JsonResponse({'gst_no': ''})
    
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from .models import Client, Branch

@login_required
@csrf_exempt
def get_client_contacts(request):
    """
    Returns all branch contacts for a given client.
    Used in Add New Lead page to populate Contact Person dropdown and autofill data.
    """
    client_id = request.GET.get('client_id')

    if not client_id:
        return JsonResponse({'error': 'Missing client_id'}, status=400)

    try:
        client = Client.objects.get(id=client_id)
        branches = Branch.objects.filter(client=client).values(
            'id',
            'branch_name',
            'contact_person',
            'email',      # ✅ Added email
            'mobile',
            'gst_no',
            'address',
        )

        # ✅ Return clean consistent structure
        return JsonResponse({
            'client': {
                'company_name': client.company_name,
                'default_email': client.email or '',
                'default_mobile': client.mobile or '',
                'default_address': client.address or '',
                'default_contact': client.contact_person or '',
                'default_gst_no': client.gst_no or '',
            },
            'branches': list(branches),
        })

    except Client.DoesNotExist:
        return JsonResponse({'error': 'Client not found'}, status=404)



from decimal import Decimal, InvalidOperation
from django.db import transaction
from django.db.models import F
from django.utils.timezone import now
from django.shortcuts import render, redirect, get_object_or_404
import re

from .models import (
    Client, Lead, Estimation, EstimationItem,
    TermsAndConditions, GSTSettings, EstimationSettings
)

# =====================================================
# DEFAULT TERMS (one per line)
# =====================================================
DEFAULT_TERMS = """Payment Terms: 100% Advance Payment or As Per Agreed Terms
Service Warranty 30 to 90 Days Depending upon the Availed Service
All Products and Accessories Carries Standard OEM Warranty"""

# =====================================================
# SAFE DECIMAL
# =====================================================
def safe_decimal(value, default='0.00'):
    try:
        return Decimal(value)
    except (InvalidOperation, TypeError, ValueError):
        return Decimal(default)

# =====================================================
# SPLIT + CLEAN TERMS
# =====================================================
_BULLET_PREFIX = re.compile(r'^\s*[-•]\s*')

def _split_lines(text: str):
    """
    Split text into clean individual lines.
    Removes bullets and empty lines.
    """
    lines = []
    for raw in (text or "").replace("\r\n", "\n").split("\n"):
        clean = _BULLET_PREFIX.sub("", raw.strip())
        if clean:
            lines.append(clean)
    return lines

# =====================================================
# MERGE + NUMBER TERMS PROPERLY
# =====================================================
def merge_terms_to_html(default_terms_text, user_terms_text):
    """
    Merge default and user terms,
    remove duplicates,
    and render as:

    1) Term
    2) Term
    3) Term
    """

    default_lines = _split_lines(default_terms_text)
    user_lines = _split_lines(user_terms_text)

    # Remove duplicates (case insensitive)
    seen = set()
    merged = []

    for line in default_lines + user_lines:
        key = line.lower()
        if key not in seen:
            seen.add(key)
            merged.append(line)

    # Generate numbered output (PDF-safe)
    return "".join(
        f"<div style='margin-bottom:4px;'>{i+1}) {line}</div>"
        for i, line in enumerate(merged)
    )



@transaction.atomic
def generate_and_reserve_quote_no():
    """
    Atomically generate and reserve the next quote number.
    """
    setting = (
        EstimationSettings.objects.select_for_update().first()
        or EstimationSettings.objects.create(prefix="EST", next_number=1)
    )
    while True:
        quote_no = f"{setting.prefix}-{setting.next_number:04d}"
        if not Estimation.objects.filter(quote_no=quote_no).exists():
            setting.next_number = F('next_number') + 1
            setting.save(update_fields=['next_number'])
            setting.refresh_from_db(fields=['next_number'])
            return quote_no
        setting.next_number = F('next_number') + 1
        setting.save(update_fields=['next_number'])
        setting.refresh_from_db(fields=['next_number'])

def create_quotation(request):
    clients = Client.objects.all().order_by('company_name')
    pending_leads = Lead.objects.filter(status='Pending').order_by('-id')

    gst_setting = GSTSettings.objects.order_by('-id').first()
    gst_percentage = float(getattr(gst_setting, 'gst_percentage', 18.0))

    # Prefer DB terms; fallback to DEFAULT_TERMS literal (plain text)
    terms_obj = TermsAndConditions.objects.order_by('-id').first()
    default_terms_text = (terms_obj.content.strip() if terms_obj and terms_obj.content else DEFAULT_TERMS)

    if request.method == 'POST':
        company_id = request.POST.get('company_name')
        lead_id = request.POST.get('lead_no') or None

        try:
            with transaction.atomic():
                quote_no = generate_and_reserve_quote_no()
                quote_date = now().date()

                client = get_object_or_404(Client, id=company_id)
                lead_instance = Lead.objects.filter(id=lead_id).first() if lead_id else None

                # Merge default and user-entered lines into HTML list
                user_terms_text = request.POST.get('terms_conditions') or ""
                final_terms_html = merge_terms_to_html(default_terms_text, user_terms_text)

                estimation = Estimation.objects.create(
                    quote_no=quote_no,
                    quote_date=quote_date,
                    company_name=client,
                    lead_no=lead_instance,
                    validity_days=int(request.POST.get('validity_days') or 0),
                    gst_no=(request.POST.get('gst_no') or "").strip(),
                    billing_address=(request.POST.get('billing_address') or "").strip(),
                    shipping_address=(request.POST.get('shipping_address') or "").strip(),
                    terms_conditions=final_terms_html,  # Store aligned HTML list
                    bank_details=(request.POST.get('bank_details') or "").strip(),
                    sub_total=safe_decimal(request.POST.get('sub_total')),
                    discount=safe_decimal(request.POST.get('discount')),
                    gst_amount=safe_decimal(request.POST.get('gst_amount')),
                    total=safe_decimal(request.POST.get('total')),
                    status='Pending',
                )

                items = zip(
                    request.POST.getlist('item_details[]'),
                    request.POST.getlist('hsn_sac[]'),
                    request.POST.getlist('quantity[]'),
                    request.POST.getlist('uom[]'),
                    request.POST.getlist('rate[]'),
                    request.POST.getlist('tax[]'),
                    request.POST.getlist('amount[]'),
                )

                for detail, hsn, qty, uom, rate, tax, amt in items:
                    detail_clean = (detail or "").strip()
                    if not detail_clean:
                        continue
                    EstimationItem.objects.create(
                        estimation=estimation,
                        item_details=detail_clean,
                        hsn_sac=((hsn or "").strip() or None),
                        quantity=int(qty or 0),
                        uom=uom or "Nos",
                        rate=safe_decimal(rate),
                        tax=safe_decimal(tax),
                        amount=safe_decimal(amt),
                    )

                # Optional: update derived lead status
                if lead_instance:
                    lead_instance.computed_status = 'Quoted'
                    lead_instance.save(update_fields=['computed_status'])

            return redirect('estimation_list')

        except Exception as e:
            return render(request, 'create_quotation.html', {
                'clients': clients,
                'pending_leads': pending_leads,
                'gst_percentage': gst_percentage,
                'terms': default_terms_text,  # keep default visible
                'error': f"Something went wrong: {e}",
            })

    # GET
    return render(request, 'create_quotation.html', {
        'clients': clients,
        'pending_leads': pending_leads,
        'gst_percentage': gst_percentage,
        'terms': default_terms_text,
    })

from django.shortcuts import get_object_or_404, render
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.views import View
from datetime import timedelta
from pathlib import Path
from decimal import Decimal   # 🔥 REQUIRED
from weasyprint import HTML
from django.conf import settings
from .models import Estimation

from .models import Estimation, TermsAndConditions
from .utils import inr_currency_words


class QuotationPDFView(View):
    def get(self, request, pk):
        estimation = get_object_or_404(Estimation, pk=pk)

        items = estimation.items.all()

        sub_total = estimation.sub_total or Decimal("0")
        discount = estimation.discount or Decimal("0")
        taxable_value = sub_total - discount

        gst_rate = Decimal("18")
        gst_amount = (taxable_value * gst_rate) / Decimal("100")

        customer_gst = (estimation.gst_no or "").strip()
        OUR_GST_STATE = getattr(settings, "GST_STATE_CODE", "29")

        same_state = not customer_gst or customer_gst[:2] == OUR_GST_STATE

        if same_state:
            cgst = sgst = gst_amount / 2
            cgst_rate = sgst_rate = gst_rate / 2
            igst = Decimal("0")
            igst_rate = Decimal("0")
        else:
            cgst = sgst = Decimal("0")
            cgst_rate = sgst_rate = Decimal("0")
            igst = gst_amount
            igst_rate = gst_rate

        total = taxable_value + gst_amount

        expiry_date = estimation.quote_date + timedelta(
            days=estimation.validity_days or 0
        )

        amount_in_words = inr_currency_words(total)

        # =========================================
        # 🔥 CONVERT UL/LI TERMS TO NUMBERED FORMAT
        # =========================================
        from bs4 import BeautifulSoup
        import re

        raw_terms = estimation.terms_conditions or ""

        # 1️⃣ Convert any HTML to plain text
        soup = BeautifulSoup(raw_terms, "html.parser")
        plain_text = soup.get_text("\n")  # force newline separation

        # 2️⃣ Remove ALL numbering patterns like:
        # "1)", "2)", "10)" anywhere in text
        plain_text = re.sub(r'\b\d+\)\s*', '', plain_text)

        # 3️⃣ Split properly on newline
        lines = plain_text.replace("\r\n", "\n").split("\n")

        # 4️⃣ Clean empty lines
        cleaned_terms = [line.strip() for line in lines if line.strip()]

        # 5️⃣ Generate fresh numbering
        numbered_terms = "".join(
            f"<div style='margin-bottom:4px;'>{i+1}) {text}</div>"
            for i, text in enumerate(cleaned_terms)
        )


        # Logo
        logo_path = Path(settings.STATIC_ROOT) / "images/logo.png"
        logo_uri = logo_path.as_uri()

        context = {
            "estimation": estimation,
            "items": items,
            "sub_total": sub_total,
            "discount": discount,
            "taxable_value": taxable_value,
            "gst_amount": gst_amount,
            "total": total,
            "same_state": same_state,
            "cgst": cgst,
            "sgst": sgst,
            "igst": igst,
            "cgst_rate": cgst_rate,
            "sgst_rate": sgst_rate,
            "igst_rate": igst_rate,
            "terms": numbered_terms,   # 🔥 IMPORTANT
            "expiry_date": expiry_date,
            "amount_in_words": amount_in_words,
            "logo_uri": logo_uri,
        }

        html_string = render_to_string(
            "quotation_pdf_template.html",
            context
        )

        pdf = HTML(
            string=html_string,
            base_url=request.build_absolute_uri("/")
        ).write_pdf()

        response = HttpResponse(pdf, content_type="application/pdf")
        response["Content-Disposition"] = (
            f'inline; filename="Quotation_{estimation.quote_no}.pdf"'
        )

        return response


    
import logging
from decimal import Decimal
from django.db import transaction
from django.shortcuts import get_object_or_404, redirect, render
from django.utils.timezone import localdate
from django.core.paginator import Paginator
from django.db.models import Q

from .models import Estimation, EstimationItem, Client, Lead, TermsAndConditions
from .forms import EstimationForm

logger = logging.getLogger(__name__)


# ====================================================
# Helper Function
# ====================================================
def _d(val, default='0.00'):
    """Safely convert value to Decimal."""
    try:
        return Decimal(val or default)
    except Exception:
        return Decimal(default)


# ====================================================
# Estimation View (List + Search + Sort)
# ====================================================
def estimation_view(request):
    """
    Displays list of estimations with optional sorting and search.
    """
    sort = request.GET.get('sort', 'quote_date')
    query = request.GET.get('q', '')

    # Sorting logic
    if sort == 'company':
        order_field = 'company_name__company_name'
    else:
        order_field = '-quote_date'

    estimations = Estimation.objects.all().order_by(order_field)

    # Search logic
    if query:
        estimations = estimations.filter(
            Q(quote_no__icontains=query) |
            Q(company_name__company_name__icontains=query)
        )

    # Pagination
    paginator = Paginator(estimations, 15)
    page_obj = paginator.get_page(request.GET.get('page'))

    return render(request, 'estimation_list.html', {
        'page_obj': page_obj,
        'query': query,
        'current_sort': sort,
    })

# Replace direct bs4 import with safe optional import and fallback
try:
    from bs4 import BeautifulSoup  # type: ignore
    _HAS_BS4 = True
except Exception:
    _HAS_BS4 = False

    # Minimal fallback using stdlib for simple <li> extraction.
    from html import unescape
    import re

    def _extract_li_texts(html_text):
        """
        Very small fallback that extracts text from <li>...</li> blocks,
        strips inner HTML tags and unescapes HTML entities.
        Suitable for simple list-html produced/stored by this app.
        """
        if not html_text:
            return []
        # capture anything between <li ...> and </li> (multiline)
        li_matches = re.findall(r'<li[^>]*>(.*?)</li>', html_text, flags=re.I | re.S)
        results = []
        for m in li_matches:
            # remove any remaining tags inside the li
            text = re.sub(r'<[^>]+>', '', m)
            text = unescape(text).strip()
            if text:
                results.append(text)
        return results


def terms_html_to_text(html):
    """
    Converts <ul><li>...</li></ul> into plain text lines.
    Uses BeautifulSoup when available, otherwise uses a safe fallback.
    """
    if not html:
        return ""
    if _HAS_BS4:
        soup = BeautifulSoup(html, "html.parser")
        return "\n".join(li.get_text(strip=True) for li in soup.find_all("li"))
    else:
        return "\n".join(_extract_li_texts(html))


# ====================================================
# Edit Estimation
# ====================================================
def edit_estimation(request, pk):
    estimation = get_object_or_404(Estimation, pk=pk)
    clients = Client.objects.all()
    items = EstimationItem.objects.filter(estimation=estimation).order_by('id')
    all_leads = Lead.objects.filter(company_name=estimation.company_name)

    # 🔹 Convert stored HTML → plain text for edit screen
    terms_plain = terms_html_to_text(estimation.terms_conditions)

    form = EstimationForm(request.POST or None, instance=estimation)

    if request.method == 'POST':
        try:
            if not form.is_valid():
                return render(request, 'edit_estimation.html', {
                    'form': form,
                    'estimation': estimation,
                    'clients': clients,
                    'items': items,
                    'all_leads': all_leads,
                    'terms_plain': terms_plain, 
                    'error': "Please fix the highlighted errors.",
                })

            with transaction.atomic():
                updated = form.save(commit=False)

                # Basic fields
                updated.company_name_id = request.POST.get('company_name') or estimation.company_name_id
                updated.lead_no_id = request.POST.get('lead_no') or None
                updated.quote_date = request.POST.get('quote_date') or updated.quote_date
                updated.validity_days = request.POST.get('validity_days') or updated.validity_days

                # Amount fields
                updated.sub_total = _d(request.POST.get('sub_total'))
                updated.discount = _d(request.POST.get('discount'))
                updated.gst_amount = _d(request.POST.get('gst_amount'))
                updated.total = _d(request.POST.get('total'))

                # 🔹 SAVE PLAIN TEXT ONLY
                updated.terms_conditions = request.POST.get('terms_conditions', '').strip()


                updated.save()

                # Replace items
                EstimationItem.objects.filter(estimation=updated).delete()

                for detail, hsn, qty, uom, rate, tax, amt in zip(
                    request.POST.getlist('item_details[]'),
                    request.POST.getlist('hsn_sac[]'),
                    request.POST.getlist('quantity[]'),
                    request.POST.getlist('uom[]'),
                    request.POST.getlist('rate[]'),
                    request.POST.getlist('tax[]'),
                    request.POST.getlist('amount[]'),
                ):
                    if detail.strip():
                        EstimationItem.objects.create(
                            estimation=updated,
                            item_details=detail.strip(),
                            hsn_sac=(hsn or "").strip() or None,
                            quantity=int(qty or 0),
                            uom=uom or "Nos",
                            rate=_d(rate),
                            tax=_d(tax),
                            amount=_d(amt),
                        )

                logger.info(f"✅ Estimation {updated.quote_no} updated successfully.")
                return redirect('estimation_list')

        except Exception as e:
            logger.exception("Error saving estimation %s", estimation.pk)
            form.add_error(None, f"Error saving quotation: {e}")

    # 🔹 GET request
    return render(request, 'edit_estimation.html', {
        'form': form,
        'estimation': estimation,
        'clients': clients,
        'items': items,
        'all_leads': all_leads,
        'terms_plain': terms_plain,   # ✅ THIS IS THE KEY
    })


from django.db.models import Sum, Count
from decimal import Decimal


# ====================================================
# Estimation List (Follow-up Filter)
# ====================================================
def estimation_list(request):
    """
    Filter estimations by today's follow-up.
    """
    today = localdate()
    follow_up_filter = request.GET.get("follow_up", "")

    if follow_up_filter == "today":
        estimations = Estimation.objects.filter(follow_up_date=today).order_by('-id')
    else:
        estimations = Estimation.objects.all().order_by('-id') 

    paginator = Paginator(estimations, 15)
    page_obj = paginator.get_page(request.GET.get("page"))

    return render(request, "estimation_list.html", {
        "page_obj": page_obj,
        "follow_up": follow_up_filter,
        "today": today,
    })

from django.shortcuts import get_object_or_404, redirect, render
from django.contrib import messages
from django.views.decorators.http import require_http_methods
from .models import Estimation, Invoice
from .utils import generate_invoice_number


@require_http_methods(["GET", "POST"])
def approve_estimation(request, pk):
    estimation = get_object_or_404(Estimation, pk=pk)

    # ==========================
    # SHOW APPROVAL FORM (GET)
    # ==========================
    if request.method == "GET":
        return render(
            request,
            "crm/approve_estimation.html",
            {"estimation": estimation}
        )

    # ==========================
    # APPROVE + GENERATE INVOICE (POST)
    # ==========================
    # Prevent duplicate invoice
    if Invoice.objects.filter(estimation=estimation).exists():
        messages.warning(request, "Invoice already generated.")
        return redirect("invoice_list")

    # Update estimation details
    estimation.credit_days = request.POST.get("credit_days") or 0
    estimation.po_number = request.POST.get("po_number")
    estimation.po_date = request.POST.get("po_date") or None
    estimation.po_received_date = request.POST.get("po_received_date") or None
    estimation.remarks = request.POST.get("remarks")

    if "po_attachment" in request.FILES:
        estimation.po_attachment = request.FILES["po_attachment"]

    # ✅ CREATE INVOICE IMMEDIATELY
    Invoice.objects.create(
        estimation=estimation,
        invoice_no=generate_invoice_number(),
        total_value=estimation.total,
        paid_amount=0,
        balance_due=estimation.total,
        credit_days=estimation.credit_days or 0,
        remarks=estimation.remarks or "",
        status="Unpaid",
        is_approved=True,   # 🔑 makes it appear in Generated Invoices
    )

    # ✅ MOVE ESTIMATION OUT OF APPROVAL FLOW
    estimation.status = "Invoiced"
    estimation.save(update_fields=["status", "credit_days", "po_number", "po_date",
                                   "po_received_date", "remarks", "po_attachment"])

    messages.success(request, "Invoice generated successfully.")

    # 🚀 GO DIRECTLY TO GENERATED INVOICES
    return redirect("invoice_list")


@require_POST
def reject_estimation(request, pk):
    estimation = get_object_or_404(Estimation, pk=pk)
    estimation.status = "Rejected"
    estimation.remarks = request.POST.get("reason", "")
    estimation.save()
    return redirect("estimation")

def invoice_approval_table(request):
    estimations = Estimation.objects.filter(
        status="Approved",
        invoices__isnull=True   # ✅ FIXED
    ).order_by("-created_at")

    invoices = Invoice.objects.all().order_by("-created_at")

    return render(
        request,
        "crm/invoice_approval_list.html",
        {
            "estimations": estimations,
            "invoices": invoices,
        }
    )



@require_POST
def reject_invoice(request, pk):
    try:
        invoice = Invoice.objects.get(pk=pk)
        estimation = invoice.estimation
        invoice.delete()
    except Invoice.DoesNotExist:
        estimation = get_object_or_404(Estimation, pk=pk)
    estimation.status = 'Rejected'
    estimation.remarks = request.POST.get('reason', '')
    estimation.save()
    return redirect('invoice_approval_list')

@csrf_exempt
def mark_as_lost(request, pk):
    estimation = get_object_or_404(Estimation, pk=pk)
    if request.method == "POST":
        reason = request.POST.get("reason", "")
        estimation.status = "Lost"
        estimation.lost_reason = reason
        estimation.save()
        return JsonResponse({"status": "success"})
    if estimation.status == "Lost":
        return JsonResponse({"reason": estimation.lost_reason})
    return JsonResponse({"status": "need_reason"})

@require_http_methods(["POST"])
def mark_lost(request, pk):
    estimation = get_object_or_404(Estimation, pk=pk)
    reason = request.POST.get('reason', '').strip()
    if reason:
        estimation.status = "Lost"
        estimation.lost_reason = reason
        estimation.save()
    return redirect('estimation_list')

def create_invoice(request):
    return render(request, 'create_invoice.html')

def update_estimation_status(request, pk, new_status):
    estimation = get_object_or_404(Estimation, pk=pk)
    if new_status == "rejected" and request.method == "POST":
        estimation.status = "Rejected"
        estimation.remarks = request.POST.get("reason", "")
    else:
        estimation.status = new_status
    estimation.save()
    return redirect("invoice_approval_list")

def generate_invoice_number():
    last = Invoice.objects.order_by('-id').first()
    number = int(last.invoice_no.split('-')[-1]) + 1 if last else 1
    return f"INV-{number:04d}"

def invoice_detail_view(request, pk):
    estimation = get_object_or_404(Estimation, pk=pk)
    items = EstimationItem.objects.filter(estimation=estimation)
    invoice = Invoice.objects.filter(estimation=estimation).first()
    return render(request, 'crm/invoice_detail.html', {
        'estimation': estimation, 'items': items, 'invoice': invoice, 'amount_in_words': inr_currency_words(estimation.total),
    })

from django.views.decorators.http import require_POST
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages

@require_POST
def approve_invoice(request, est_id):
    estimation = get_object_or_404(Estimation, id=est_id)
    action = request.POST.get("action")

    # ❌ Reject
    if action == "reject":
        estimation.status = "Rejected"
        estimation.remarks = request.POST.get("reason", "")
        estimation.save()
        messages.error(request, "Quotation rejected.")
        return redirect("invoice_list")

    # ✅ Approve
    if action == "approve":
        if Invoice.objects.filter(estimation=estimation).exists():
            messages.warning(request, "Invoice already exists.")
            return redirect("invoice_list")

        Invoice.objects.create(
            estimation=estimation,
            invoice_no=generate_invoice_number(),
            total_value=estimation.total,
            paid_amount=0,
            balance_due=estimation.total,
            credit_days=estimation.credit_days or 0,
            remarks=estimation.remarks or "",
            status="Unpaid",
            is_approved=True,   # 🔑 THIS WAS MISSING
        )

        estimation.status = "Invoiced"
        estimation.save(update_fields=["status"])

        messages.success(request, "Invoice generated successfully.")
        return redirect("invoice_list")

    return redirect("invoice_list")



from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages

def generate_invoice_from_estimation(request, pk):
    estimation = get_object_or_404(Estimation, pk=pk)

    # 🔒 Prevent duplicate invoice
    if Invoice.objects.filter(estimation=estimation).exists():
        messages.warning(request, "Invoice already exists.")
        return redirect("invoice_list")

    # ✅ Create invoice correctly
    Invoice.objects.create(
        estimation=estimation,
        invoice_no=generate_invoice_number(),
        total_value=estimation.total,
        paid_amount=0,
        balance_due=estimation.total,
        credit_days=estimation.credit_days or 0,
        remarks=estimation.remarks or "",
        status="Unpaid",
        is_approved=True,   # 🔑 CRITICAL
    )

    # ✅ Move estimation out of approvals
    estimation.status = "Invoiced"
    estimation.save(update_fields=["status"])

    messages.success(request, "Invoice generated successfully.")

    # ✅ GO TO GENERATED INVOICES TABLE
    return redirect("invoice_list")

def invoice_list(request):
    invoices = Invoice.objects.filter(
        is_approved=True
    ).order_by("-created_at")

    return render(
        request,
        "crm/invoice_list.html",
        {"invoices": invoices}
    )

from decimal import Decimal

def estimation_detail_view(request, pk):
    estimation = get_object_or_404(Estimation, pk=pk)
    items = estimation.items.all()
    total = sum((item.amount for item in items), Decimal("0.00"))

    return render(
        request,
        "crm/estimation_detail.html",
        {
            "estimation": estimation,
            "items": items,
            "total": total,
        }
    )



from django.shortcuts import get_object_or_404, redirect
from django.views.decorators.http import require_POST
from django.contrib import messages

@require_POST
def mark_under_review(request, pk):
    est = get_object_or_404(Estimation, pk=pk)

    est.status = "Pending"  # ✅ valid status
    est.follow_up_date = request.POST.get("follow_up_date")
    est.follow_up_remarks = request.POST.get("follow_up_remarks")
    est.save()

    messages.success(request, "Estimation marked for follow-up")
    return redirect("estimation_list")


def invoices_view(request):
    estimations_without_invoice = (
        Estimation.objects
        .filter(invoices__isnull=True)   # ✅ CORRECT
        .exclude(status="Rejected")
    )

    return render(
        request,
        "crm/invoice_approval_list.html",
        {
            "estimations_without_invoice": estimations_without_invoice
        }
    )



from decimal import Decimal
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from weasyprint import HTML
from datetime import timedelta
from pathlib import Path
from django.conf import settings
from num2words import num2words
import re


def invoice_pdf_view(request, invoice_id):
    invoice = get_object_or_404(Invoice, pk=invoice_id)
    invoice.refresh_from_db()

    estimation = invoice.estimation
    items = estimation.items.all()

    # =========================
    # DUE DATE
    # =========================
    due_date = invoice.invoice_date + timedelta(days=invoice.credit_days or 0)

   # =========================
    # ITEM CALCULATION (CORRECT)
    # =========================
    sub_total = Decimal("0.00")
    gst_amount = Decimal("0.00")

    for item in items:
        qty = Decimal(item.quantity or 0)
        rate = Decimal(item.rate or 0)
        tax = Decimal(item.tax or 0)

        base = qty * rate
        tax_amt = (base * tax) / Decimal("100")

        item.base_amount = base.quantize(Decimal("0.01"))
        item.tax_amount = tax_amt.quantize(Decimal("0.01"))

        sub_total += base
        gst_amount += tax_amt

    sub_total = sub_total.quantize(Decimal("0.01"))
    gst_amount = gst_amount.quantize(Decimal("0.01"))
    total = (sub_total + gst_amount).quantize(Decimal("0.01"))



    # =========================
    # AMOUNT IN WORDS
    # =========================
    rupees = int(total)
    paise = int((total - rupees) * 100)

    amount_in_words = f"Rupees {num2words(rupees, lang='en_IN').title()}"
    if paise > 0:
        amount_in_words += f" and {num2words(paise, lang='en_IN').title()} Paise"
    amount_in_words += " Only"

    # =========================
    # GST SPLIT (DISPLAY)
    # =========================
    customer_gst = (estimation.gst_no or "").strip()
    our_state_code = "29"

    same_state = customer_gst[:2] == our_state_code if customer_gst else True

    if same_state:
        cgst = (gst_amount / 2).quantize(Decimal("0.01"))
        sgst = (gst_amount / 2).quantize(Decimal("0.01"))
        igst = Decimal("0.00")
    else:
        cgst = Decimal("0.00")
        sgst = Decimal("0.00")
        igst = gst_amount


    # =========================
    # CLEAN TERMS FOR PDF
    # =========================
    raw_terms = estimation.terms_conditions or ""

    # Convert HTML → clean text
    plain_terms = strip_tags(raw_terms)
    plain_terms = re.sub(r'^\s*\d+\)\s*', '', plain_terms, flags=re.MULTILINE)

    terms_pdf = [
        line.strip()
        for line in plain_terms.split("\n")
        if line.strip()
    ]

    # =========================
    # LOGO
    # =========================
    logo_path = Path(settings.STATIC_ROOT) / "images/logo.png"

    # =========================
    # RENDER PDF
    # =========================
    html_string = render_to_string(
        "invoice_pdf_weasy.html",
        {
            "invoice": invoice,
            "estimation": estimation,
            "items": items,
            "due_date": due_date,
            "sub_total": sub_total,
            "gst_amount": gst_amount,
            "total": total,
            "amount_in_words": amount_in_words,
            "same_state": same_state,
            "cgst": cgst,
            "sgst": sgst,
            "igst": igst,
            "terms_pdf": terms_pdf,
            "logo_uri": logo_path.as_uri(),
        }
    )

    pdf = HTML(
        string=html_string,
        base_url=settings.STATIC_ROOT.as_uri()
    ).write_pdf()

    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="{invoice.invoice_no}.pdf"'
    response["Cache-Control"] = "no-store"

    return response



from decimal import Decimal
from django.shortcuts import get_object_or_404, render, redirect
from django.forms import modelformset_factory
from django.utils.dateparse import parse_date
from django.db import transaction
from django.contrib import messages

from .models import Invoice, EstimationItem


def edit_invoice(request, invoice_id):
    invoice = get_object_or_404(Invoice, pk=invoice_id)
    estimation = invoice.estimation

    # 🔒 Block paid invoices
    if not invoice.can_edit:
        messages.error(request, "Paid invoices cannot be edited.")
        return redirect("invoice_list")

    ItemFormSet = modelformset_factory(
        EstimationItem,
        fields=("item_details", "hsn_sac", "quantity", "uom", "rate", "tax"),
        extra=0,
        can_delete=True
    )

    if request.method == "POST":
        formset = ItemFormSet(request.POST, queryset=estimation.items.all())

        if not formset.is_valid():
            messages.error(request, "Please fix item errors.")
            return render(request, "crm/edit_invoice.html", {
                "invoice": invoice,
                "estimation": estimation,
                "formset": formset,
            })

        with transaction.atomic():

            # =========================
            # 1️⃣ Invoice Header
            # =========================
            invoice_date = request.POST.get("invoice_date")
            if invoice_date:
                invoice.invoice_date = parse_date(invoice_date)

            invoice.credit_days = int(request.POST.get("credit_days") or 0)
            invoice.save(update_fields=["invoice_date", "credit_days"])

            # =========================
            # 2️⃣ Estimation Header
            # =========================
            estimation.po_number = request.POST.get("po_number")
            estimation.billing_address = request.POST.get("billing_address")
            estimation.shipping_address = request.POST.get("shipping_address")
            estimation.gst_no = request.POST.get("gst_no", "")
            estimation.terms_conditions = request.POST.get("terms_conditions", "")
            estimation.save()

            # =========================
            # 3️⃣ Items Calculation
            # =========================
            sub_total = Decimal("0.00")
            gst_total = Decimal("0.00")

            items = formset.save(commit=False)

            # Delete removed items
            for obj in formset.deleted_objects:
                obj.delete()

            for item in items:
                item.estimation = estimation

                base = item.quantity * item.rate
                tax_amt = (base * item.tax) / Decimal("100")

                item.amount = base + tax_amt
                item.save()

                sub_total += base
                gst_total += tax_amt

            total = sub_total + gst_total

            # =========================
            # 4️⃣ Prevent Negative Balance
            # =========================
            if total < invoice.paid_amount:
                messages.error(
                    request,
                    f"Total cannot be less than Paid Amount (₹{invoice.paid_amount})."
                )
                return redirect("edit_invoice", invoice_id=invoice.id)

            # =========================
            # 5️⃣ Save Totals
            # =========================
            estimation.sub_total = sub_total
            estimation.gst_amount = gst_total
            estimation.total = total
            estimation.save(update_fields=["sub_total", "gst_amount", "total"])

            invoice.total_value = total
            invoice.balance_due = total - invoice.paid_amount
            invoice.save(update_fields=["total_value", "balance_due"])

        messages.success(request, "Invoice updated successfully")
        return redirect("invoice_list")

    # =========================
    # GET REQUEST
    # =========================
    formset = ItemFormSet(queryset=estimation.items.all())

    return render(request, "crm/edit_invoice.html", {
        "invoice": invoice,
        "estimation": estimation,
        "formset": formset,
    })

def recalculate_paid_amount(self):
    from django.db.models import Sum
    total_paid = self.logs.aggregate(
        total=Sum("amount_paid")
    )["total"] or Decimal("0.00")

    self.paid_amount = total_paid

    balance = self.total_value - total_paid
    if balance < 0:
        balance = Decimal("0.00")

    self.balance_due = balance
    self.save(update_fields=["paid_amount", "balance_due"])


@require_POST
def update_payment_status(request, pk):
    invoice = get_object_or_404(Invoice, pk=pk)
    new_status = request.POST.get("payment_status")
    if new_status in dict(Invoice.PAYMENT_STATUS_CHOICES):
        invoice.payment_status = new_status
        invoice.save()
    return redirect('invoice_approval_table')

from decimal import Decimal
from django.shortcuts import get_object_or_404, redirect
from .models import Invoice, PaymentLog

def confirm_payment_post(request, invoice_id):
    invoice = get_object_or_404(Invoice, pk=invoice_id)

    try:
        amount_paid = Decimal(request.POST.get("amount_paid", "0"))
        utr_number = request.POST.get("utr_number")
        payment_date = request.POST.get("payment_date")

        # ✅ Create payment log (money received)
        PaymentLog.objects.create(
            invoice=invoice,
            amount_paid=amount_paid,
            utr_number=utr_number,
            payment_date=payment_date,
            status="Partial Paid"
        )

        # ✅ SINGLE SOURCE OF TRUTH
        invoice.recalculate_paid_amount()

    except Exception as e:
        return HttpResponse(f"Something went wrong: {e}")

    return redirect("invoice_list")
def get_payment_logs(request, invoice_id):
    invoice = Invoice.objects.get(pk=invoice_id)
    logs = invoice.logs.all().order_by('-payment_date')
    logs_data = [
        {"amount_paid": str(log.amount_paid), "utr_number": log.utr_number, "payment_date": log.payment_date.strftime('%Y-%m-%d')}
        for log in logs
    ]
    return JsonResponse({"logs": logs_data})

def view_payment_logs(request, invoice_id):
    invoice = get_object_or_404(Invoice, id=invoice_id)
    logs = PaymentLog.objects.filter(invoice=invoice).order_by('-payment_date')
    return render(request, 'payment_logs.html', {'invoice': invoice, 'logs': logs})

from datetime import date
from django.db.models import Sum
from django.http import HttpResponse
import openpyxl

def invoice_list_view(request):
    invoices = Invoice.objects.all()

    # ---- Date Filters ----
    filter_type = request.GET.get("range", "month")
    start_date = end_date = None

    today = date.today()

    if filter_type == "month":
        start_date = today.replace(day=1)
        end_date = today

    elif filter_type == "fy":
        if today.month >= 4:
            start_date = date(today.year, 4, 1)
            end_date = date(today.year + 1, 3, 31)
        else:
            start_date = date(today.year - 1, 4, 1)
            end_date = date(today.year, 3, 31)

    elif filter_type == "custom":
        start_date = request.GET.get("start_date")
        end_date = request.GET.get("end_date")

    if start_date and end_date:
        invoices = invoices.filter(created_at__date__range=[start_date, end_date])

    invoices = invoices.order_by("-created_at")

    from django.db.models import Sum
    from .models import PaymentLog

    summary = invoices.aggregate(
        total_amount=Sum("total_value"),
        balance_amount=Sum("balance_due"),
    )

    # ✅ Paid Amount = ALL money received (NO STATUS FILTER)
    paid_amount = PaymentLog.objects.filter(
        invoice__in=invoices
    ).aggregate(
        total=Sum("amount_paid")
    )["total"] or 0

    gst_summary = invoices.aggregate(
        gst_collected=Sum("estimation__gst_amount")
    )

    context = {
        "invoices": invoices,
        "summary": {
            "count": invoices.count(),
            "total": summary["total_amount"] or 0,
            "paid": paid_amount,                      # ✅ FIXED
            "balance": summary["balance_amount"] or 0,
            "gst": gst_summary["gst_collected"] or 0,
        },
        "filter_type": filter_type,
        "start_date": start_date,
        "end_date": end_date,
    }

    return render(request, "crm/invoice_approval_list.html", context)


# crm/views.py
import openpyxl
from django.http import HttpResponse
from datetime import date
from .models import Invoice

def export_invoice_summary(request):
    invoices = Invoice.objects.select_related("estimation", "estimation__company_name")

    # 🔹 Date filters (same as invoice list)
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    if start_date and end_date:
        invoices = invoices.filter(created_at__date__range=[start_date, end_date])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice Summary"

    # 🔹 Header row
    ws.append([
        "Invoice No",
        "Invoice Date",
        "Client",
        "Total Amount",
        "Paid Amount",
        "Balance Due",
        "Status",
    ])

    # 🔹 Data rows
    for inv in invoices:
        ws.append([
            inv.invoice_no,
            inv.created_at.strftime("%d-%m-%Y"),
            str(inv.estimation.company_name),
            float(inv.total_value),
            float(inv.paid_amount),
            float(inv.balance_due),
            inv.status,
        ])

    # 🔹 Response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        "attachment; filename=invoice_summary.xlsx"
    )

    wb.save(response)
    return response


def invoice_logs_api(request, invoice_id):
    logs = PaymentLog.objects.filter(invoice_id=invoice_id).order_by('-payment_date')
    data = [{
        'amount_paid': str(log.amount_paid),
        'payment_date': log.payment_date.strftime('%d-%m-%Y'),
        'utr_number': log.utr_number
    } for log in logs]
    return JsonResponse(data, safe=False)

@login_required
def purchase_order_view(request):
    return render(request, 'purchase_order.html')

@login_required
def vendor_view(request):
    return render(request, 'vendor.html')

@login_required
def bill_view(request):
    return render(request, 'bill.html')

@login_required
def profile_view(request):
    return render(request, 'profile.html')

@login_required
def report_list(request):
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    leads = Lead.objects.all()
    if from_date and to_date:
        leads = leads.filter(date__range=[from_date, to_date])
    else:
        leads = leads.order_by('-date')
    report_data = []
    for lead in leads:
        estimation = Estimation.objects.filter(lead_no=lead).first()
        invoice = Invoice.objects.filter(estimation__lead_no=lead).first()
        report_data.append({
            'lead_no': lead.lead_no,
            'lead_date': lead.date.strftime('%d-%m-%Y') if lead.date else '',
            'client': lead.company_name.company_name if lead.company_name else '',
            'requirement': lead.requirement,
            'estimation_no': estimation.quote_no if estimation else '',
            'estimation_status': estimation.status if estimation else '',
            'lost_reason': estimation.lost_reason if estimation else '',
            'po_number': estimation.po_number if estimation else '',
            'po_date': estimation.po_date.strftime('%d-%m-%Y') if estimation and estimation.po_date else '',
            'po_attachment': estimation.po_attachment.url if estimation and estimation.po_attachment else '',
            'invoice_no': invoice.invoice_no if invoice else '',
            'invoice_amount': invoice.total_value if invoice else '',
            'payment_status': invoice.status if invoice else '',
        })
    paginator = Paginator(report_data, 20)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'crm/report_list.html', {'report_data': page_obj, 'page_obj': page_obj})

def export_report_excel(request):
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    leads = Lead.objects.all()
    if from_date and to_date:
        leads = leads.filter(date__range=[from_date, to_date])
    data = []
    for lead in leads:
        estimation = Estimation.objects.filter(lead_no=lead.lead_no).first()
        invoice = Invoice.objects.filter(estimation__lead_no=lead.lead_no).first()
        data.append({
            "Lead No": lead.lead_no,
            "Lead Date": lead.date.strftime('%d-%m-%Y') if lead.date else '',
            "Client": lead.company_name.company_name,
            "Requirement": lead.requirement,
            "Estimation No": estimation.quote_no if estimation else '',
            "Estimation Status": estimation.status if estimation else 'Pending',
            "Lost Reason": estimation.lost_reason if estimation else '',
            "Client PO Number": estimation.po_number if estimation else '',
            "PO Date": estimation.po_date.strftime("%d-%m-%Y") if estimation and estimation.po_date else '',
            "Invoice No": invoice.invoice_no if invoice else '',
            "Amount": float(invoice.total_value) if invoice else '',
            "Paid": float(invoice.total_value) - float(invoice.balance_due) if invoice else '',
            "Balance": float(invoice.balance_due) if invoice else '',
            "Payment Status": invoice.status if invoice else '',
        })
    import pandas as pd
    df = pd.DataFrame(data)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=CRM_Complete_Report.xlsx'
    df.to_excel(response, index=False)
    return response

def export_report_pdf(request):
    invoices = get_filtered_invoices(request)
    html_string = render_to_string('report_pdf_template.html', {'invoices': invoices})
    from weasyprint import HTML
    pdf_file = HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf()
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Filtered_CRM_Report.pdf"'
    return response

def get_filtered_invoices(request):
    invoices = Invoice.objects.select_related('estimation', 'estimation__company_name').all()


    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    company = request.GET.get('company')
    lead_no = request.GET.get('lead_no')

    if from_date and to_date:
        invoices = invoices.filter(created_at__date__range=[from_date, to_date])
    elif from_date:
        invoices = invoices.filter(created_at__date__gte=from_date)
    elif to_date:
        invoices = invoices.filter(created_at__date__lte=to_date)

    if company:
        invoices = invoices.filter(estimation__company_name__id=company)
    if lead_no:
        invoices = invoices.filter(estimation__lead_no=lead_no)

    return invoices

import openpyxl
from django.http import HttpResponse
from decimal import Decimal
from .models import Invoice
from django.conf import settings


def export_gst_excel(request):
    invoices = Invoice.objects.select_related(
        "estimation", "estimation__company_name"
    )

    # ---- Date filters ----
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    if start_date and end_date:
        invoices = invoices.filter(created_at__date__range=[start_date, end_date])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GST Report"

    # ---- Header (GST-ready) ----
    ws.append([
        "Invoice No",
        "Invoice Date",
        "Client Name",
        "Client GSTIN",
        "State Type",
        "Taxable Value",
        "CGST",
        "SGST",
        "IGST",
        "Total GST",
        "Invoice Total",
    ])

    OUR_STATE_CODE = getattr(settings, "GST_STATE_CODE", "29")  # Karnataka default

    for inv in invoices:
        est = inv.estimation
        gst_no = est.gst_no or ""

        taxable_value = est.sub_total - est.discount
        gst_amount = est.gst_amount

        # ---- GST Logic ----
        if gst_no and gst_no[:2] == OUR_STATE_CODE:
            cgst = gst_amount / 2
            sgst = gst_amount / 2
            igst = Decimal("0.00")
            state_type = "Intra-State"
        else:
            cgst = Decimal("0.00")
            sgst = Decimal("0.00")
            igst = gst_amount
            state_type = "Inter-State"

        ws.append([
            inv.invoice_no,
            inv.created_at.strftime("%d-%m-%Y"),
            str(est.company_name),
            gst_no or "URP",
            state_type,
            float(taxable_value),
            float(cgst),
            float(sgst),
            float(igst),
            float(gst_amount),
            float(inv.total_value),
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=GST_Report.xlsx"

    wb.save(response)
    return response



from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, redirect, render
from django.utils.timezone import now

from .models import Estimation, DeliveryChallan, DeliveryChallanItem, EstimationItem


def generate_dc_number():
    """
    Generate a new Delivery Challan number in the form DC-0001, DC-0002, ...
    Falls back gracefully if the last dc_no isn't parseable.
    """
    last = DeliveryChallan.objects.order_by('-id').first()
    if last and getattr(last, 'dc_no', None):
        try:
            number = int(str(last.dc_no).split('-')[-1]) + 1
        except Exception:
            # fallback to using last.id + 1 to avoid crash
            number = (last.id or 0) + 1
    else:
        number = 1
    return f"DC-{number:04d}"


from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, redirect, render
from django.utils.timezone import now

@login_required
def create_dc(request, pk):
    estimation = get_object_or_404(
        Estimation.objects.prefetch_related("items"),
        pk=pk
    )

    if request.method == "POST":
        selected_items = request.POST.getlist("item_id[]")

        if not selected_items:
            messages.error(request, "Please select at least one item.")
            return redirect("create_dc", pk=pk)

        # ✅ Create Delivery Challan FIRST
        dc = DeliveryChallan.objects.create(
            estimation=estimation,
            dc_no=generate_dc_number(),
            dc_date=request.POST.get("dc_date"),
            delivery_address=request.POST.get("delivery_address"),
            contact_person=request.POST.get("contact_person"),
            contact_number=request.POST.get("contact_number"),
            terms=request.POST.get("terms", "ABC"),
            po_no=request.POST.get("po_no"),
            po_date=request.POST.get("po_date") or None,
        )

        try:
            for item_id in selected_items:
                qty = int(request.POST.get(f"qty_{item_id}", 0))
                desc = request.POST.get(f"desc_{item_id}", "").strip()
                uom = request.POST.get(f"uom_{item_id}", "Nos").strip()

                if qty <= 0:
                    continue

                est_item = get_object_or_404(EstimationItem, id=item_id)

                # ❌ Prevent excess quantity
                if qty > est_item.quantity:
                    raise ValueError(
                        f"Quantity exceeds limit for {est_item.item_details}"
                    )

                DeliveryChallanItem.objects.create(
                    dc=dc,
                    estimation_item=est_item,
                    quantity=qty,
                    uom=uom,
                    description=desc or est_item.item_details,
                )

        except Exception as e:
            dc.delete()  # ✅ rollback DC
            messages.error(request, str(e))
            return redirect("create_dc", pk=pk)

        messages.success(
            request,
            f"Delivery Challan {dc.dc_no} created successfully."
        )
        return redirect("dc_list")

    # ---------------- GET REQUEST ----------------
    lead = getattr(estimation, "lead_no", None)

    context = {
        "estimation": estimation,
        "dc_no": generate_dc_number(),
        "dc_date": now().date(),

        # From quotation
        "delivery_address": estimation.shipping_address or "",

        # From lead
        "contact_person": lead.contact_person if lead else "",
        "contact_number": lead.mobile if lead else "",

        "items": estimation.items.all(),
        "terms": (
            "Received the above mentioned goods in good condition, complaints (if any) contact within 24 hours."
        ),
    }

    return render(request, "dc/create_dc.html", context)

@login_required
def dc_list(request):
    dcs = DeliveryChallan.objects.select_related(
        'estimation'
    ).order_by('-created_at')

    return render(request, 'dc/dc_list.html', {'dcs': dcs})

from pathlib import Path
from django.conf import settings
from django.http import HttpResponse
from django.template.loader import render_to_string
from weasyprint import HTML
from django.shortcuts import get_object_or_404
from django.contrib.auth.decorators import login_required

@login_required
def dc_pdf(request, pk):
    dc = get_object_or_404(DeliveryChallan, pk=pk)

    # Resolve logo path safely — STATIC_ROOT may be None in some environments
    logo_uri = ""
    try:
        if getattr(settings, "STATIC_ROOT", None):
            logo_path = Path(settings.STATIC_ROOT) / "images" / "logo.png"
            logo_uri = logo_path.resolve().as_uri()
        else:
            # Fallback to BASE_DIR/static/images/logo.png if STATIC_ROOT not set
            logo_path = Path(settings.BASE_DIR) / "static" / "images" / "logo.png"
            logo_uri = logo_path.resolve().as_uri()
    except Exception:
        # keep logo_uri empty string if resolution fails
        logo_uri = ""

    context = {
        "dc": dc,
        "company_name": "iSecure Solutions",
        "company_address": "#60 Swarupa, 5th West Cr, Riches Garden, RM Nagar, Bangalore 560016",
        "company_phone": "916 916 8216",
        "company_email": "support@isecuresolutions.in",
        "company_gstin": "29AVXPP2341P1ZJ",
        "logo_uri": logo_uri,
    }

    html_string = render_to_string("dc/dc_pdf.html", context)

    # Use a reliable base_url for weasyprint (convert Path to URI)
    try:
        base_url = Path(settings.BASE_DIR).resolve().as_uri()
    except Exception:
        base_url = request.build_absolute_uri('/')

    pdf = HTML(
        string=html_string,
        base_url=base_url
    ).write_pdf()

    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="{dc.dc_no}.pdf"'
    return response



@login_required
@require_POST
def delete_dc(request, pk):
    dc = get_object_or_404(DeliveryChallan, pk=pk)
    dc.delete()
    messages.success(request, "Delivery Challan deleted successfully.")
    return redirect("dc_list")


from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.db import transaction
from django.contrib.auth.decorators import login_required

@login_required
def edit_dc(request, pk):
    dc = get_object_or_404(DeliveryChallan, pk=pk)

    # 🔒 Lock if invoice exists
    if Invoice.objects.filter(estimation=dc.estimation).exists():
        messages.warning(
            request,
            "⚠️ This Delivery Challan cannot be edited because an invoice exists."
        )
        return redirect('dc_list')

    items = DeliveryChallanItem.objects.filter(dc=dc).select_related('estimation_item')

    if request.method == "POST":
        try:
            with transaction.atomic():

                # ===== UPDATE HEADER =====
                dc.dc_date = request.POST.get('dc_date') or dc.dc_date
                dc.delivery_address = request.POST.get('delivery_address', '').strip()
                dc.contact_person = request.POST.get('contact_person', '').strip()
                dc.contact_number = request.POST.get('contact_number', '').strip()
                dc.po_no = request.POST.get('po_no', '').strip()
                dc.po_date = request.POST.get('po_date') or None
                dc.terms = request.POST.get('terms', '').strip()
                dc.save()

                # ===== REPLACE ITEMS =====
                DeliveryChallanItem.objects.filter(dc=dc).delete()

                est_item_ids = request.POST.getlist('estimation_item_id[]')
                quantities = request.POST.getlist('quantity[]')
                uoms = request.POST.getlist('uom[]')
                descriptions = request.POST.getlist('description[]')

                for est_id, qty, uom, desc in zip(
                    est_item_ids, quantities, uoms, descriptions
                ):
                    if not qty or int(qty) <= 0:
                        continue

                    est_item = EstimationItem.objects.get(pk=est_id)

                    DeliveryChallanItem.objects.create(
                        dc=dc,
                        estimation_item=est_item,
                        quantity=int(qty),
                        uom=uom,
                        description=desc.strip() or est_item.item_details
                    )

                messages.success(request, "✅ Delivery Challan updated successfully.")
                return redirect('dc_list')

        except Exception as e:
            messages.error(request, f"❌ Error updating DC: {e}")

    return render(request, 'crm/edit_dc.html', {
        'dc': dc,
        'items': items,
    })

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json
from .models import CallLog


@csrf_exempt
def save_call_log(request):

    if request.method != "POST":
        return JsonResponse({"error": "Invalid request"}, status=400)

    try:
        data = json.loads(request.body)
    except Exception:
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    phone = data.get("phone_number")
    call_type = data.get("call_type")
    duration = data.get("duration", 0)
    sim_slot = data.get("sim_slot", "")
    name = data.get("name")
    address = data.get("address", "")
    remarks = data.get("remarks")
    status_value = data.get("status")

    if not phone:
        return JsonResponse({"error": "Phone number required"}, status=400)

    # 🔥 Always get latest record for that number
    existing = CallLog.objects.filter(
        phone_number=phone
    ).order_by("-call_time").first()

    # ==========================
    # IF FORM SAVE (name or remarks present)
    # ==========================
    if existing and (name or remarks):

        existing.name = name or existing.name
        existing.address = address or existing.address
        existing.remarks = remarks or existing.remarks
        existing.status = status_value or existing.status
        existing.save()

        return JsonResponse({"status": "updated"})

    # ==========================
    # OTHERWISE CREATE (receiver auto save)
    # ==========================
    CallLog.objects.create(
        phone_number=phone,
        name=name,
        address=address,
        remarks=remarks,
        status=status_value or "follow_up",
        call_type=call_type,
        duration=duration,
        sim_slot=sim_slot
    )

    return JsonResponse({"status": "created"})

 
# ===============================
# CALL LOG LIST PAGE
# ===============================
from django.shortcuts import render
from django.db.models import Q
from django.utils import timezone
from datetime import datetime, timedelta

from .models import CallLog

def call_log_list(request):
    """
    List call logs with advanced filtering:
    q - general text (phone, name, remarks, address)
    status - follow_up/junk/lead_stage/existing_client/missed
    call_type - incoming/missed
    sim_slot - SIM 1 / SIM 2 / ...
    start_date, end_date - YYYY-MM-DD
    min_duration, max_duration - integer seconds
    """
    qs = CallLog.objects.all().order_by("-call_time")

    q = request.GET.get("q", "").strip()
    status = request.GET.get("status", "").strip()
    call_type = request.GET.get("call_type", "").strip()
    sim_slot = request.GET.get("sim_slot", "").strip()
    start_date = request.GET.get("start_date", "").strip()
    end_date = request.GET.get("end_date", "").strip()
    min_duration = request.GET.get("min_duration", "").strip()
    max_duration = request.GET.get("max_duration", "").strip()

    # general search
    if q:
        qs = qs.filter(
            Q(phone_number__icontains=q) |
            Q(name__icontains=q) |
            Q(remarks__icontains=q) |
            Q(address__icontains=q)
        )

    if status:
        qs = qs.filter(status=status)

    if call_type:
        qs = qs.filter(call_type=call_type)

    if sim_slot:
        # store sim slot normalized in DB like "SIM 1", "SIM 2" - adjust as needed
        qs = qs.filter(sim_slot__iexact=sim_slot)

    # date range parsing (safe)
    try:
        if start_date:
            start_dt = datetime.strptime(start_date, "%Y-%m-%d")
            start_dt = timezone.make_aware(datetime.combine(start_dt.date(), datetime.min.time()))
            qs = qs.filter(call_time__gte=start_dt)
        if end_date:
            end_dt = datetime.strptime(end_date, "%Y-%m-%d")
            # include whole day
            end_dt = timezone.make_aware(datetime.combine(end_dt.date(), datetime.max.time()))
            qs = qs.filter(call_time__lte=end_dt)
    except Exception:
        # ignore parse errors (or add messages)
        pass

    # duration range
    try:
        if min_duration:
            qs = qs.filter(duration__gte=int(min_duration))
        if max_duration:
            qs = qs.filter(duration__lte=int(max_duration))
    except ValueError:
        pass

    # KPI counts for the filtered set (so KPIs update when filters applied)
    context = {
        "logs": qs,  # consider pagination later
        "query": q,
        "total_calls": qs.count(),
        "incoming_calls": qs.filter(call_type="incoming").count(),
        "missed_calls": qs.filter(call_type="missed").count(),
        "lead_stage": qs.filter(status="lead_stage").count(),

        # echo current filters back to template
        "filter_status": status,
        "filter_call_type": call_type,
        "filter_sim_slot": sim_slot,
        "filter_start_date": start_date,
        "filter_end_date": end_date,
        "filter_min_duration": min_duration,
        "filter_max_duration": max_duration,
    }

    return render(request, "crm/call_log_list.html", context)


# ===============================
# UPDATE STATUS INLINE
# ===============================
def update_call_status(request, pk):
    if request.method == "POST":
        log = get_object_or_404(CallLog, pk=pk)
        log.status = request.POST.get("status")
        log.save()

    return redirect("call_log_list")


# ===============================
# EXPORT TO EXCEL
# ===============================
def export_call_logs(request):

    # 🔥 IMPORTANT: DO NOT FILTER BY USER (You are not saving user in API)
    logs = CallLog.objects.all().order_by("-call_time")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Call Logs"

    ws.append([
        "Date & Time",
        "Phone",
        "Name",
        "Address",
        "Remarks",
        "Status",
        "Type",
        "Duration",
        "SIM Slot"
    ])

    for log in logs:
        ws.append([
            log.call_time.strftime("%d-%m-%Y %H:%M"),
            log.phone_number,
            log.name,
            log.address,
            log.remarks,
            log.get_status_display(),
            log.get_call_type_display(),
            f"{log.duration} sec"
            if log.duration else "",
            log.sim_slot or "", 
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=call_logs.xlsx"

    wb.save(response)
    return response

from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
import json
from .models import CallLog


@csrf_exempt
def update_call_log_field(request, pk):

    if request.method != "POST":
        return JsonResponse({"error": "Invalid method"}, status=400)

    try:
        data = json.loads(request.body)
    except Exception:
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    field = data.get("field")
    value = data.get("value")

    if not field:
        return JsonResponse({"error": "Field missing"}, status=400)

    try:
        log = CallLog.objects.get(pk=pk)
    except CallLog.DoesNotExist:
        return JsonResponse({"error": "Log not found"}, status=404)

    # 🔥 Allowed editable fields only
    allowed_fields = ["name", "address", "remarks", "status"]

    if field not in allowed_fields:
        return JsonResponse({"error": "Field not allowed"}, status=400)

    setattr(log, field, value)
    log.save()

    return JsonResponse({"status": "success"})